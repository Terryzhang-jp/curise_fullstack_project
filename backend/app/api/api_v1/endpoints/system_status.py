from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, inspect, func
from typing import Dict, List, Any, Optional
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import urllib.parse
from datetime import datetime

from app.api.deps import get_db, get_current_active_user
from app.models.models import (
    User, Country, Port, Company, Ship, Category,
    Supplier, Product, Order, OrderItem, CruiseOrder,
    EmailTemplate
)

router = APIRouter()
logger = logging.getLogger(__name__)

@router.get("/system-status")
async def get_system_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """获取系统所有表的数据状态"""
    
    try:
        # 定义所有表的配置
        table_configs = [
            {
                "name": "用户管理",
                "table": "users", 
                "model": User,
                "icon": "👥",
                "description": "系统用户账户",
                "priority": 1
            },
            {
                "name": "国家数据",
                "table": "countries",
                "model": Country, 
                "icon": "🌍",
                "description": "国家和地区信息",
                "priority": 2
            },
            {
                "name": "港口数据", 
                "table": "ports",
                "model": Port,
                "icon": "🚢",
                "description": "港口和码头信息",
                "priority": 3
            },
            {
                "name": "公司数据",
                "table": "companies", 
                "model": Company,
                "icon": "🏢",
                "description": "邮轮公司信息",
                "priority": 3
            },
            {
                "name": "船舶数据",
                "table": "ships",
                "model": Ship,
                "icon": "⛵",
                "description": "邮轮船舶信息", 
                "priority": 4
            },
            {
                "name": "产品类别",
                "table": "categories",
                "model": Category,
                "icon": "📂",
                "description": "产品分类信息",
                "priority": 2
            },
            {
                "name": "供应商数据",
                "table": "suppliers",
                "model": Supplier,
                "icon": "🏪",
                "description": "供应商信息",
                "priority": 3
            },
            {
                "name": "产品数据",
                "table": "products", 
                "model": Product,
                "icon": "📦",
                "description": "产品和商品信息",
                "priority": 4
            },
            {
                "name": "订单数据",
                "table": "orders",
                "model": Order,
                "icon": "📋",
                "description": "采购订单信息",
                "priority": 5
            },
            {
                "name": "邮轮订单",
                "table": "cruise_orders", 
                "model": CruiseOrder,
                "icon": "🛳️",
                "description": "邮轮专用订单",
                "priority": 5
            }
        ]
        
        # 获取每个表的记录数
        table_status = []
        total_records = 0
        empty_tables = 0
        
        for config in table_configs:
            try:
                count = db.query(config["model"]).count()
                total_records += count
                
                if count == 0:
                    empty_tables += 1
                
                status = "empty" if count == 0 else "has_data"
                
                table_status.append({
                    "name": config["name"],
                    "table": config["table"],
                    "icon": config["icon"],
                    "description": config["description"],
                    "count": count,
                    "status": status,
                    "priority": config["priority"]
                })
                
            except Exception as e:
                logger.error(f"获取表 {config['table']} 数据时出错: {e}")
                table_status.append({
                    "name": config["name"],
                    "table": config["table"], 
                    "icon": config["icon"],
                    "description": config["description"],
                    "count": 0,
                    "status": "error",
                    "priority": config["priority"],
                    "error": str(e)
                })
        
        # 按优先级排序
        table_status.sort(key=lambda x: x["priority"])
        
        # 计算系统状态
        system_health = "excellent" if empty_tables == 0 else \
                       "good" if empty_tables <= 3 else \
                       "needs_attention" if empty_tables <= 6 else "poor"
        
        # 生成建议
        recommendations = []
        if empty_tables > 0:
            recommendations.append({
                "type": "warning",
                "message": f"发现 {empty_tables} 个表没有数据，建议先导入基础数据"
            })
        
        if total_records < 50:
            recommendations.append({
                "type": "info", 
                "message": "系统数据较少，建议导入更多数据以获得更好的使用体验"
            })
        
        if total_records == 0:
            recommendations.append({
                "type": "urgent",
                "message": "系统暂无任何数据，强烈建议先进行数据导入"
            })
        
        return {
            "system_health": system_health,
            "total_records": total_records,
            "empty_tables": empty_tables,
            "total_tables": len(table_status),
            "tables": table_status,
            "recommendations": recommendations,
            "last_check": "now"
        }

    except Exception as e:
        logger.error(f"获取系统状态时出错: {e}")
        raise HTTPException(status_code=500, detail=f"获取系统状态失败: {str(e)}")

@router.get("/data-quality-analysis")
async def get_data_quality_analysis(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """获取数据质量分析报告"""

    try:
        # 定义需要分析的表配置
        table_configs = [
            {
                "name": "产品数据",
                "table": "products",
                "model": Product,
                "icon": "📦",
                "key_fields": ["product_name_en", "country_id", "category_id", "effective_from"],
                "foreign_keys": {
                    "country_id": {"table": "countries", "model": Country},
                    "category_id": {"table": "categories", "model": Category},
                    "supplier_id": {"table": "suppliers", "model": Supplier},
                    "port_id": {"table": "ports", "model": Port}
                }
            },
            {
                "name": "供应商数据",
                "table": "suppliers",
                "model": Supplier,
                "icon": "🏪",
                "key_fields": ["name", "country_id"],
                "foreign_keys": {
                    "country_id": {"table": "countries", "model": Country}
                }
            },
            {
                "name": "港口数据",
                "table": "ports",
                "model": Port,
                "icon": "🚢",
                "key_fields": ["name", "country_id"],
                "foreign_keys": {
                    "country_id": {"table": "countries", "model": Country}
                }
            },
            {
                "name": "公司数据",
                "table": "companies",
                "model": Company,
                "icon": "🏢",
                "key_fields": ["name", "country_id"],
                "foreign_keys": {
                    "country_id": {"table": "countries", "model": Country}
                }
            },
            {
                "name": "船舶数据",
                "table": "ships",
                "model": Ship,
                "icon": "⛵",
                "key_fields": ["name", "company_id"],
                "foreign_keys": {
                    "company_id": {"table": "companies", "model": Company}
                }
            }
        ]

        analysis_results = []

        for config in table_configs:
            try:
                # 获取表的基本信息
                total_count = db.query(config["model"]).count()

                if total_count == 0:
                    analysis_results.append({
                        "table_name": config["name"],
                        "table": config["table"],
                        "icon": config["icon"],
                        "total_records": 0,
                        "missing_rate_analysis": {},
                        "foreign_key_integrity": {},
                        "quality_score": 0,
                        "status": "empty"
                    })
                    continue

                # 1. 字段缺失率统计
                missing_rate_analysis = await analyze_missing_rates(db, config, total_count)

                # 2. 外键关系完整性检查
                foreign_key_integrity = await analyze_foreign_key_integrity(db, config, total_count)

                # 3. 计算表级数据质量评分
                quality_score = calculate_quality_score(missing_rate_analysis, foreign_key_integrity)

                analysis_results.append({
                    "table_name": config["name"],
                    "table": config["table"],
                    "icon": config["icon"],
                    "total_records": total_count,
                    "missing_rate_analysis": missing_rate_analysis,
                    "foreign_key_integrity": foreign_key_integrity,
                    "quality_score": quality_score,
                    "status": "analyzed"
                })

            except Exception as e:
                logger.error(f"分析表 {config['table']} 时出错: {e}")
                analysis_results.append({
                    "table_name": config["name"],
                    "table": config["table"],
                    "icon": config["icon"],
                    "total_records": 0,
                    "missing_rate_analysis": {},
                    "foreign_key_integrity": {},
                    "quality_score": 0,
                    "status": "error",
                    "error": str(e)
                })

        # 计算整体数据质量评分
        overall_score = calculate_overall_quality_score(analysis_results)

        return {
            "overall_quality_score": overall_score,
            "analysis_timestamp": datetime.now().isoformat(),
            "tables": analysis_results
        }

    except Exception as e:
        logger.error(f"数据质量分析时出错: {e}")
        raise HTTPException(status_code=500, detail=f"数据质量分析失败: {str(e)}")

@router.get("/table-templates")
async def get_table_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """获取所有表的导入模板信息"""
    
    templates = [
        {
            "table": "countries",
            "name": "国家数据",
            "icon": "🌍", 
            "description": "国家和地区基础信息",
            "columns": ["name", "code", "status"],
            "required_columns": ["name", "code"],
            "example_data": [
                {"name": "中国", "code": "CN", "status": "true"},
                {"name": "日本", "code": "JP", "status": "true"}
            ],
            "dependencies": [],
            "priority": 1
        },
        {
            "table": "categories", 
            "name": "产品类别",
            "icon": "📂",
            "description": "产品分类信息",
            "columns": ["name", "code", "description", "status"],
            "required_columns": ["name"],
            "example_data": [
                {"name": "食品饮料", "code": "FOOD", "description": "食品和饮料类产品", "status": "true"},
                {"name": "日用品", "code": "DAILY", "description": "日常生活用品", "status": "true"}
            ],
            "dependencies": [],
            "priority": 1
        },
        {
            "table": "ports",
            "name": "港口数据",
            "icon": "🚢",
            "description": "港口和码头信息",
            "columns": ["name", "code", "country_name", "location", "status"],
            "required_columns": ["name", "country_name"],
            "example_data": [
                {"name": "上海港", "code": "CNSHA", "country_name": "中国", "location": "上海市", "status": "true"},
                {"name": "横滨港", "code": "JPYOK", "country_name": "日本", "location": "横滨市", "status": "true"}
            ],
            "dependencies": ["countries"],
            "priority": 2
        },
        {
            "table": "companies",
            "name": "公司数据",
            "icon": "🏢",
            "description": "邮轮公司信息",
            "columns": ["name", "country_name", "contact", "email", "phone", "status"],
            "required_columns": ["name", "country_name"],
            "example_data": [
                {"name": "皇家加勒比", "country_name": "美国", "contact": "John Smith", "email": "contact@rccl.com", "phone": "+1-123-456-7890", "status": "true"}
            ],
            "dependencies": ["countries"],
            "priority": 2
        },
        {
            "table": "suppliers",
            "name": "供应商数据",
            "icon": "🏪",
            "description": "供应商信息",
            "columns": ["name", "country_name", "contact", "email", "phone", "status"],
            "required_columns": ["name", "country_name"],
            "example_data": [
                {"name": "ABC食品供应商", "country_name": "中国", "contact": "张三", "email": "zhang@abc.com", "phone": "+86-138-0000-0000", "status": "true"}
            ],
            "dependencies": ["countries"],
            "priority": 2
        },
        {
            "table": "ships",
            "name": "船舶数据",
            "icon": "⛵",
            "description": "邮轮船舶信息",
            "columns": ["name", "company_name", "ship_type", "capacity", "status"],
            "required_columns": ["name", "company_name", "capacity"],
            "example_data": [
                {"name": "海洋魅力号", "company_name": "皇家加勒比", "ship_type": "大型邮轮", "capacity": "3000", "status": "true"}
            ],
            "dependencies": ["companies"],
            "priority": 3
        },
        {
            "table": "products",
            "name": "产品数据", 
            "icon": "📦",
            "description": "产品和商品信息",
            "columns": ["product_name_en", "product_name_jp", "code", "country_name", "category_name", "supplier_name", "port_name", "unit", "price", "currency", "unit_size", "pack_size", "brand", "country_of_origin", "effective_from", "effective_to"],
            "required_columns": ["product_name_en", "country_name", "category_name", "effective_from"],
            "example_data": [
                {"product_name_en": "Apple Juice", "product_name_jp": "アップルジュース", "code": "JUICE001", "country_name": "中国", "category_name": "饮料", "supplier_name": "ABC食品供应商", "port_name": "上海港", "unit": "瓶", "price": "5.50", "currency": "CNY", "unit_size": "500ml", "pack_size": "24", "brand": "ABC", "country_of_origin": "中国", "effective_from": "2025-01-01", "effective_to": "2025-06-01"},
                {"product_name_en": "Orange Juice", "product_name_jp": "オレンジジュース", "code": "JUICE002", "country_name": "日本", "category_name": "饮料", "supplier_name": "BAC食品供应公司", "port_name": "神户港", "unit": "瓶", "price": "6.00", "currency": "JPY", "unit_size": "500ml", "pack_size": "12", "brand": "XYZ", "country_of_origin": "日本", "effective_from": "2025-02-01", "effective_to": ""}
            ],
            "dependencies": ["countries", "categories", "suppliers", "ports"],
            "priority": 4
        }
    ]
    
    return {"templates": templates}

@router.get("/download-template/{table_name}")
async def download_template(
    table_name: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """下载指定表的Excel导入模板"""

    # 获取模板配置
    templates_response = await get_table_templates(db, current_user)
    templates = templates_response["templates"]

    # 找到对应的模板
    template = None
    for t in templates:
        if t["table"] == table_name:
            template = t
            break

    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    # 创建Excel工作簿
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = template["name"]

    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # 写入标题行
    for col_idx, column in enumerate(template["columns"], 1):
        cell = worksheet.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

        # 标记必填字段
        if column in template["required_columns"]:
            cell.value = f"{column} *"

    # 写入示例数据
    for row_idx, example in enumerate(template["example_data"], 2):
        for col_idx, column in enumerate(template["columns"], 1):
            value = example.get(column, "")
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # 调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 添加说明工作表
    info_sheet = workbook.create_sheet("使用说明")
    info_data = [
        ["字段名", "是否必填", "说明"],
        ["", "", ""],
    ]

    # 字段说明映射
    field_descriptions = {
        "product_name_en": "产品名称(英文) - 产品的英文名称",
        "product_name_jp": "产品名称(日文) - 产品的日文名称",
        "code": "产品代码 - 产品的唯一标识代码",
        "country_name": "国家名称 - 产品所属国家，必须在系统中已存在",
        "category_name": "类别名称 - 产品类别，必须在系统中已存在",
        "supplier_name": "供应商名称 - 产品供应商，必须在系统中已存在",
        "port_name": "港口名称 - 产品相关港口，必须在系统中已存在",
        "unit": "单位 - 产品计量单位",
        "price": "价格 - 产品价格，数字格式",
        "currency": "货币 - 价格货币单位",
        "unit_size": "规格 - 产品规格描述",
        "pack_size": "包装数量 - 每包装的数量，数字格式",
        "brand": "品牌 - 产品品牌名称",
        "country_of_origin": "原产国 - 产品原产国",
        "effective_from": "起始日期 - 产品有效期开始日期，格式：YYYY-MM-DD，必填",
        "effective_to": "结束日期 - 产品有效期结束日期，格式：YYYY-MM-DD，可选，为空时自动设置为起始日期+3个月"
    }

    for column in template["columns"]:
        is_required = "是" if column in template["required_columns"] else "否"
        description = field_descriptions.get(column, f"{column}字段")
        info_data.append([column, is_required, description])

    # 添加依赖说明
    if template["dependencies"]:
        info_data.extend([
            ["", "", ""],
            ["依赖关系说明:", "", ""],
            [f"此表依赖以下数据表: {', '.join(template['dependencies'])}", "", ""],
            ["请确保先导入依赖的数据表", "", ""]
        ])

    # 添加特别说明（针对产品表）
    if template["table"] == "products":
        info_data.extend([
            ["", "", ""],
            ["重要说明:", "", ""],
            ["1. 起始日期(effective_from)为必填字段", "", ""],
            ["2. 结束日期(effective_to)可选，为空时自动设置为起始日期+3个月", "", ""],
            ["3. 日期格式支持: YYYY-MM-DD 或 YYYY/MM/DD", "", ""],
            ["4. 国家、类别、供应商、港口名称必须在系统中已存在", "", ""],
            ["5. 价格和包装数量请使用数字格式", "", ""]
        ])

    for row_idx, row_data in enumerate(info_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = info_sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # 标题行
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

    # 调整说明表列宽
    for column in info_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        info_sheet.column_dimensions[column_letter].width = adjusted_width

    # 保存到内存
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # 返回文件 - 使用URL编码处理中文文件名
    filename = f"{template['name']}_import_template.xlsx"
    encoded_filename = urllib.parse.quote(filename.encode('utf-8'))

    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

async def analyze_missing_rates(db: Session, config: Dict, total_count: int) -> Dict[str, Any]:
    """分析字段缺失率"""

    missing_analysis = {}
    model = config["model"]

    # 获取表的所有列
    inspector = inspect(db.bind)
    columns = inspector.get_columns(config["table"])

    for column in columns:
        column_name = column["name"]

        # 跳过系统字段
        if column_name in ["id", "created_at", "updated_at"]:
            continue

        try:
            # 计算非空记录数
            non_null_count = db.query(model).filter(
                getattr(model, column_name).isnot(None)
            ).count()

            # 对于字符串字段，还要检查空字符串
            if hasattr(column["type"], "python_type") and column["type"].python_type == str:
                non_empty_count = db.query(model).filter(
                    getattr(model, column_name).isnot(None),
                    getattr(model, column_name) != ""
                ).count()
                missing_count = total_count - non_empty_count
            else:
                missing_count = total_count - non_null_count

            missing_rate = (missing_count / total_count * 100) if total_count > 0 else 0

            # 判断字段重要性
            is_key_field = column_name in config.get("key_fields", [])
            is_foreign_key = column_name in config.get("foreign_keys", {})

            missing_analysis[column_name] = {
                "missing_count": missing_count,
                "total_count": total_count,
                "missing_rate": round(missing_rate, 2),
                "is_key_field": is_key_field,
                "is_foreign_key": is_foreign_key,
                "status": "critical" if missing_rate > 50 and is_key_field else
                         "warning" if missing_rate > 20 else
                         "good" if missing_rate < 5 else "acceptable"
            }

        except Exception as e:
            logger.error(f"分析字段 {column_name} 缺失率时出错: {e}")
            missing_analysis[column_name] = {
                "missing_count": 0,
                "total_count": total_count,
                "missing_rate": 0,
                "is_key_field": False,
                "is_foreign_key": False,
                "status": "error",
                "error": str(e)
            }

    return missing_analysis

async def analyze_foreign_key_integrity(db: Session, config: Dict, total_count: int) -> Dict[str, Any]:
    """分析外键关系完整性"""

    integrity_analysis = {}
    model = config["model"]
    foreign_keys = config.get("foreign_keys", {})

    for fk_column, fk_config in foreign_keys.items():
        try:
            # 检查外键完整性
            # 查找有外键值但对应记录不存在的情况
            orphaned_query = db.query(model).filter(
                getattr(model, fk_column).isnot(None)
            ).outerjoin(
                fk_config["model"],
                getattr(model, fk_column) == fk_config["model"].id
            ).filter(
                fk_config["model"].id.is_(None)
            )

            orphaned_count = orphaned_query.count()

            # 计算有外键值的记录总数
            with_fk_count = db.query(model).filter(
                getattr(model, fk_column).isnot(None)
            ).count()

            integrity_rate = ((with_fk_count - orphaned_count) / with_fk_count * 100) if with_fk_count > 0 else 100

            integrity_analysis[fk_column] = {
                "target_table": fk_config["table"],
                "total_with_fk": with_fk_count,
                "orphaned_count": orphaned_count,
                "integrity_rate": round(integrity_rate, 2),
                "status": "critical" if integrity_rate < 90 else
                         "warning" if integrity_rate < 98 else "good"
            }

        except Exception as e:
            logger.error(f"分析外键 {fk_column} 完整性时出错: {e}")
            integrity_analysis[fk_column] = {
                "target_table": fk_config["table"],
                "total_with_fk": 0,
                "orphaned_count": 0,
                "integrity_rate": 0,
                "status": "error",
                "error": str(e)
            }

    return integrity_analysis

def calculate_quality_score(missing_analysis: Dict, integrity_analysis: Dict) -> int:
    """计算表级数据质量评分 (0-100)"""

    if not missing_analysis and not integrity_analysis:
        return 0

    total_score = 0
    weight_sum = 0

    # 字段缺失率评分 (权重60%)
    if missing_analysis:
        missing_scores = []
        for field, data in missing_analysis.items():
            if data.get("status") == "error":
                continue

            missing_rate = data["missing_rate"]
            is_key = data["is_key_field"]

            # 关键字段权重更高
            field_weight = 2 if is_key else 1

            # 根据缺失率计算分数
            if missing_rate == 0:
                field_score = 100
            elif missing_rate < 5:
                field_score = 90
            elif missing_rate < 20:
                field_score = 70
            elif missing_rate < 50:
                field_score = 40
            else:
                field_score = 10

            weighted_score = field_score * field_weight
            missing_scores.append(weighted_score)
            weight_sum += field_weight

        if missing_scores and weight_sum > 0:
            missing_avg = sum(missing_scores) / weight_sum
            missing_component = missing_avg * 0.6
            total_score += missing_component

    # 外键完整性评分 (权重40%)
    if integrity_analysis:
        integrity_scores = []
        for fk, data in integrity_analysis.items():
            if data.get("status") == "error":
                continue

            integrity_rate = data["integrity_rate"]

            # 根据完整性率计算分数
            if integrity_rate >= 99:
                integrity_score = 100
            elif integrity_rate >= 95:
                integrity_score = 90
            elif integrity_rate >= 90:
                integrity_score = 70
            elif integrity_rate >= 80:
                integrity_score = 50
            else:
                integrity_score = 20

            integrity_scores.append(integrity_score)

        if integrity_scores:
            integrity_avg = sum(integrity_scores) / len(integrity_scores)
            integrity_component = integrity_avg * 0.4
            total_score += integrity_component

    # 如果只有一种分析，调整权重
    if not missing_analysis:
        total_score = total_score / 0.4  # 只有外键分析
    elif not integrity_analysis:
        total_score = total_score / 0.6  # 只有缺失率分析

    final_score = min(100, max(0, round(total_score)))

    return final_score

def calculate_overall_quality_score(analysis_results: List[Dict]) -> int:
    """计算整体数据质量评分"""

    valid_scores = [
        result["quality_score"]
        for result in analysis_results
        if result["status"] == "analyzed" and result["total_records"] > 0
    ]

    if not valid_scores:
        return 0

    return round(sum(valid_scores) / len(valid_scores))
