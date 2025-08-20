"""
Excel生成API v2 - 邮轮订单询价单生成
严格按照PURCHASE ORDER格式生成Excel文件
"""

from typing import List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session
from pydantic import BaseModel
from datetime import datetime, timedelta
import pandas as pd
import io
import logging
import urllib.parse
import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Border, Alignment, PatternFill

from app.api import deps
from app.models.models import Supplier, Port
# PDF相关导入暂时移除

# 导入新的Excel生成器 - 暂时注释掉，因为模块不存在
# import sys
# backend_path = os.path.join(os.path.dirname(__file__), '..', '..', '..', '..')
# sys.path.insert(0, backend_path)
# from new_excel_generator import NewExcelGenerator

# 设置日志
logger = logging.getLogger(__name__)
router = APIRouter()

def safe_write_cell(worksheet, cell_address, value):
    """
    安全地写入单元格，处理合并单元格的情况
    """
    try:
        cell = worksheet[cell_address]
        if isinstance(cell, MergedCell):
            # 如果是合并单元格，找到合并区域的主单元格
            for merged_range in worksheet.merged_cells.ranges:
                if cell_address in merged_range:
                    # 获取合并区域的左上角单元格（主单元格）
                    main_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
                    main_cell.value = value
                    logger.debug(f"📝 写入合并单元格主单元格 {main_cell.coordinate}: {value}")
                    return
            # 如果没找到合并区域，记录警告但不报错
            logger.warning(f"⚠️  无法找到合并单元格 {cell_address} 的主单元格，跳过写入")
        else:
            # 普通单元格，直接写入
            cell.value = value
            logger.debug(f"📝 写入普通单元格 {cell_address}: {value}")
    except Exception as e:
        logger.error(f"❌ 写入单元格 {cell_address} 失败: {e}")
        # 不抛出异常，继续处理其他单元格

# 数据模型
class ProductItem(BaseModel):
    """产品项目"""
    po_number: str
    product_code: str
    product_name_en: str
    product_name_jp: str
    pack_size: str  # 包装规格，显示在F列（与G列合并）
    quantity: int
    unit: str
    unit_price: float
    amount: float
    currency: str = "JPY"

class PurchaseOrderRequest(BaseModel):
    """采购订单请求"""
    supplier_id: int
    supplier_name: str
    products: List[ProductItem]
    delivery_date: str
    delivery_address: str
    total_amount: float
    currency: str = "JPY"
    invoice_number: str = ""
    voyage_number: str = ""

class ExcelPreviewData(BaseModel):
    """Excel预览数据"""
    supplier_info: Dict[str, Any]
    order_info: Dict[str, Any]
    delivery_info: Dict[str, Any]
    products: List[ProductItem]
    totals: Dict[str, Any]

class ExcelUpdateRequest(BaseModel):
    """Excel更新请求"""
    supplier_info: Dict[str, Any]
    order_info: Dict[str, Any]
    delivery_info: Dict[str, Any]
    products: List[ProductItem]

def create_purchase_order_excel_new(request: PurchaseOrderRequest, supplier: Supplier, db: Session) -> bytes:
    """
    使用新的Excel生成器创建采购订单Excel文件
    完全程序化生成，支持动态产品数量和白色背景
    """
    try:
        logger.info(f"📋 开始生成Excel文件，产品数量: {len(request.products)}")

        # 准备供应商信息
        supplier_info = {
            'name': supplier.name or 'タカナシ販売株式会社　横浜営業所',
            'postal_code': '〒224-0042',
            'address': '横浜市都筑区大熊町１０５－２',
            'tel': 'TEL:045-472-9171',
            'fax': 'FAX: 0445-472-6365'
        }

        # 准备订单信息
        order_info = {
            'date': request.delivery_date or datetime.now().strftime('%Y-%m-%d'),
            'invoice': request.invoice_number or f"{datetime.now().strftime('%Y%m%d')}-01 ML",
            'voyage': request.voyage_number or 'ML-1017'
        }

        # 准备交货信息
        delivery_info = {
            'date': request.delivery_date or datetime.now().strftime('%Y-%m-%d'),
            'address': request.delivery_address or '神奈川県横浜市中区海岸通り1-1-4'
        }

        # 转换产品数据格式
        products = []
        for i, product in enumerate(request.products):
            products.append({
                'po_number': product.po_number,
                'code': product.product_code,
                'name_en': product.product_name_en,
                'name_jp': product.product_name_jp,
                'description': product.pack_size,
                'quantity': product.quantity,
                'unit': product.unit,
                'price': product.unit_price,
                'currency': product.currency
            })

        logger.info("✅ 数据准备完成")

        # 回退到使用模板生成器（因为NewExcelGenerator模块不存在）
        logger.warning("NewExcelGenerator不可用，回退到模板生成器")
        return create_purchase_order_excel(request, supplier, db)

    except Exception as e:
        logger.error(f"Excel生成失败: {str(e)}")
        raise Exception(f"Excel生成失败: {str(e)}")

def create_purchase_order_excel(request: PurchaseOrderRequest, supplier: Supplier, db: Session) -> bytes:
    """
    基于模板创建采购订单Excel文件
    """
    try:
        # 模板文件路径
        template_path = "purchase_order_template.xlsx"

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"模板文件不存在: {template_path}")

        # 加载模板工作簿
        workbook = load_workbook(template_path)
        worksheet = workbook.active

        logger.info(f"✅ 成功加载模板文件: {template_path}")

        # 🔧 更新工作表名为当前日期格式：{MMDD}発注書 (ML)
        current_date = datetime.now()
        worksheet_name = f"{current_date.strftime('%m%d')}発注書 (ML)"
        worksheet.title = worksheet_name
        logger.info(f"📋 更新工作表名: {worksheet_name}")

        # 🔧 填充供应商信息到模板 (A4-A8区域)
        if supplier:
            safe_write_cell(worksheet, 'A4', supplier.name or "")  # 供应商名称
            # A5, A6 暂时留空 (postal_code, address)
            safe_write_cell(worksheet, 'A7', f"TEL:{supplier.phone}" if supplier.phone else "")  # 电话
            # A8 传真暂时留空
            logger.info(f"🏢 填充供应商信息: {supplier.name}")

        # 🔧 填充动态数据到模板

        # 1. 填充日期信息
        safe_write_cell(worksheet, 'K4', current_date.strftime('%Y-%m-%d'))  # DATE
        safe_write_cell(worksheet, 'K5', request.invoice_number or f"{current_date.strftime('%Y%m%d')}-02 ML")  # Invoice
        safe_write_cell(worksheet, 'K6', request.voyage_number or "ML-1017")  # Voyage

        # 2. 填充交货信息
        # 🔧 处理日期格式：统一转换为YYYY-MM-DD格式
        delivery_date_formatted = request.delivery_date
        if '/' in request.delivery_date:
            # 转换 2025/7/15 格式为 2025-07-15
            try:
                date_parts = request.delivery_date.split('/')
                if len(date_parts) == 3:
                    year, month, day = date_parts
                    delivery_date_formatted = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except:
                delivery_date_formatted = request.delivery_date

        safe_write_cell(worksheet, 'H8', delivery_date_formatted)  # Delivery Date

        # 🔧 交货地址处理：根据用户要求暂时留空白
        if request.delivery_address:
            safe_write_cell(worksheet, 'H9', request.delivery_address)  # Delivery Address
        else:
            # 清空模板中的默认地址
            safe_write_cell(worksheet, 'H9', "")

        logger.info(f"📅 填充日期信息: {current_date.strftime('%Y-%m-%d')}")
        logger.info(f"🚚 填充交货地址: {request.delivery_address}")

        # 2.5. 填充供应商信息（左侧区域）
        # 根据supplier_id从数据库获取供应商详细信息
        if supplier and supplier.name:
            safe_write_cell(worksheet, 'A4', supplier.name)  # 供应商名称

            # 如果有联系信息，填充到相应位置
            if supplier.phone:
                safe_write_cell(worksheet, 'A7', f"TEL:{supplier.phone}")

            # 如果有邮箱，可以填充到A8位置（替换原来的FAX）
            if supplier.email:
                safe_write_cell(worksheet, 'A8', f"E-mail:{supplier.email}")

            # 如果有联系人信息
            if supplier.contact:
                safe_write_cell(worksheet, 'A5', f"担当者：{supplier.contact}")

            logger.info(f"🏢 填充供应商信息: {supplier.name}")
        else:
            logger.warning(f"⚠️  供应商信息不完整，使用默认模板信息")

        # 3. 动态扩展产品行（按照用户的"模板+脚本"思路）
        template_product_rows = 10  # 模板预设的产品行数（第22-31行）
        first_product_row = 22  # 第一个产品行（样式行）
        last_template_row = first_product_row + template_product_rows - 1  # 第31行

        # 🎯 关键：总计区域的固定位置
        TOTALS_START_ROW = 33  # Sub Total 在第33行

        product_count = len(request.products)
        logger.info(f"📊 产品数量: {product_count}, 模板预设行数: {template_product_rows}")
        logger.info(f"🎯 总计区域起始行: {TOTALS_START_ROW}")

        # 3.1 如果产品数量超过模板行数，在总计区域前插入行
        if product_count > template_product_rows:
            additional_rows_needed = product_count - template_product_rows

            logger.info(f"🔧 需要插入 {additional_rows_needed} 行，在总计区域前插入（第{TOTALS_START_ROW}行前）")

            # 🚀 关键改进：在总计区域前插入，让整个总计区块下移
            worksheet.insert_rows(TOTALS_START_ROW, amount=additional_rows_needed)
            logger.info(f"✅ 在第{TOTALS_START_ROW}行前插入了{additional_rows_needed}行")

            # 🔧 修复合并单元格问题：手动移除可能影响产品数据区域的合并单元格
            logger.info(f"🔧 检查并修复合并单元格冲突")

            # 获取所有合并单元格
            merged_ranges_to_remove = []
            for merged_range in list(worksheet.merged_cells.ranges):
                # 检查是否与新插入的产品行冲突（第33-38行）
                if (merged_range.min_row <= TOTALS_START_ROW + additional_rows_needed - 1 and
                    merged_range.max_row >= TOTALS_START_ROW):
                    # 如果合并单元格影响产品数据区域，需要处理
                    logger.warning(f"⚠️  发现冲突的合并单元格: {merged_range}")
                    merged_ranges_to_remove.append(merged_range)

            # 移除冲突的合并单元格
            for merged_range in merged_ranges_to_remove:
                try:
                    worksheet.unmerge_cells(str(merged_range))
                    logger.info(f"🔧 移除冲突的合并单元格: {merged_range}")
                except Exception as e:
                    logger.warning(f"⚠️  移除合并单元格失败 {merged_range}: {e}")

            # 🔧 关键修复：为新插入的行创建F:G合并单元格
            logger.info(f"🔧 为新插入的行创建F:G合并单元格")
            for i in range(additional_rows_needed):
                new_row = TOTALS_START_ROW + i  # 第32, 33, 34, 35, 36行
                try:
                    worksheet.merge_cells(f'F{new_row}:G{new_row}')
                    logger.debug(f"✅ 成功合并 F{new_row}:G{new_row}")
                except Exception as e:
                    logger.warning(f"⚠️  合并失败 F{new_row}:G{new_row}: {e}")

            # 使用第一个产品行作为样式模板
            template_row = first_product_row  # 第22行作为样式模板

            # 注意：我们先插入行，稍后在填充数据后再应用样式
            logger.info(f"📝 行插入完成，已创建合并单元格，稍后将应用样式")

        # 3.2 清空所有产品数据区域（包括新插入的行）
        total_product_rows = max(template_product_rows, product_count)
        logger.info(f"🧹 清空产品数据区域: 第{first_product_row}-{first_product_row + total_product_rows - 1}行")
        for row in range(first_product_row, first_product_row + total_product_rows):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L']:
                safe_write_cell(worksheet, f'{col}{row}', None)

        # 4. 填充产品数据
        logger.info(f"🔧 开始填充 {product_count} 个产品，从第{first_product_row}行开始")
        for i, product in enumerate(request.products):
            row = first_product_row + i  # 从第22行开始
            logger.debug(f"📦 填充产品 {i+1} 到第{row}行: {product.product_name_en}")

            # 使用安全写入函数填充产品数据
            safe_write_cell(worksheet, f'A{row}', i + 1)  # NO.
            safe_write_cell(worksheet, f'B{row}', product.po_number)  # PO No.
            safe_write_cell(worksheet, f'C{row}', product.product_code)  # 商品コード
            safe_write_cell(worksheet, f'D{row}', product.product_name_en)  # 英語表記
            safe_write_cell(worksheet, f'E{row}', product.product_name_jp)  # 日本語表記
            safe_write_cell(worksheet, f'F{row}', product.pack_size)  # 包装规格（F列与G列合并）
            safe_write_cell(worksheet, f'H{row}', product.quantity)  # Quantity
            safe_write_cell(worksheet, f'I{row}', product.unit)  # Unit
            safe_write_cell(worksheet, f'J{row}', product.unit_price)  # Unit price
            safe_write_cell(worksheet, f'K{row}', product.currency)  # Currency
            # 🔧 添加金额计算公式
            safe_write_cell(worksheet, f'L{row}', f"=H{row}*J{row}")  # Amount公式

        logger.info(f"✅ 成功填充 {product_count} 个产品到模板（动态扩展了 {max(0, product_count - template_product_rows)} 行）")
        logger.info(f"📊 产品填充范围: 第{first_product_row}行 - 第{first_product_row + product_count - 1}行")

        # 🎨 统一设置所有产品行的白色背景
        logger.info(f"🎨 统一设置所有产品行的白色背景")
        white_fill = PatternFill(
            fill_type="solid",
            start_color="FFFFFFFF",
            end_color="FFFFFFFF"
        )

        # 🔧 修复：确保所有产品行都有白色背景，包括新插入的行
        # 计算实际的产品行范围
        if product_count <= template_product_rows:
            # 产品数量不超过模板行数，直接设置第22-31行
            product_rows_to_style = range(first_product_row, first_product_row + product_count)
            logger.info(f"📊 设置模板范围内的产品行背景色: 第{first_product_row}-{first_product_row + product_count - 1}行")
        else:
            # 产品数量超过模板行数，需要设置模板行+新插入的行
            # 模板行：第22-31行（10行）
            # 新插入的行：第32行开始的额外行
            all_product_rows = []

            # 添加模板行范围
            all_product_rows.extend(range(first_product_row, first_product_row + template_product_rows))

            # 添加新插入的行范围（第32行开始）
            additional_rows_start = first_product_row + template_product_rows  # 第32行
            additional_rows_count = product_count - template_product_rows
            all_product_rows.extend(range(additional_rows_start, additional_rows_start + additional_rows_count))

            product_rows_to_style = all_product_rows
            logger.info(f"📊 设置扩展后的产品行背景色: 模板行第{first_product_row}-{first_product_row + template_product_rows - 1}行 + 新增行第{additional_rows_start}-{additional_rows_start + additional_rows_count - 1}行")

        # 应用白色背景到所有产品行
        for row in product_rows_to_style:
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                try:
                    cell = worksheet[f'{col}{row}']
                    cell.fill = white_fill
                except Exception as e:
                    logger.warning(f"⚠️  设置{col}{row}背景色失败: {e}")

        logger.info(f"✅ 成功设置 {len(product_rows_to_style)} 行产品数据的白色背景")

        # 4.5 在填充数据后，为新插入的行应用样式（只复制样式，不覆盖数据）
        if product_count > template_product_rows:
            logger.info(f"🎨 为新插入的行应用样式（保留数据）")
            template_row = first_product_row  # 第22行作为样式模板

            for i in range(additional_rows_needed):
                new_row = TOTALS_START_ROW + i  # 第33, 34, 35, 36, 37, 38行
                logger.debug(f"🎨 为第{new_row}行应用样式（保留数据）")

                # 只复制样式，不覆盖数据
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                    try:
                        template_cell = worksheet[f'{col}{template_row}']
                        new_cell = worksheet[f'{col}{new_row}']

                        # 保存当前单元格的值
                        current_value = new_cell.value

                        # 复制样式（不包括值）
                        if hasattr(template_cell, 'font') and template_cell.font:
                            new_cell.font = Font(
                                name=template_cell.font.name,
                                size=template_cell.font.size,
                                bold=template_cell.font.bold,
                                italic=template_cell.font.italic,
                                vertAlign=template_cell.font.vertAlign,
                                underline=template_cell.font.underline,
                                strike=template_cell.font.strike,
                                color=template_cell.font.color
                            )

                        if hasattr(template_cell, 'border') and template_cell.border:
                            new_cell.border = Border(
                                left=copy(template_cell.border.left) if template_cell.border.left else None,
                                right=copy(template_cell.border.right) if template_cell.border.right else None,
                                top=copy(template_cell.border.top) if template_cell.border.top else None,
                                bottom=copy(template_cell.border.bottom) if template_cell.border.bottom else None,
                                diagonal=copy(template_cell.border.diagonal) if template_cell.border.diagonal else None,
                                diagonal_direction=template_cell.border.diagonal_direction,
                                outline=template_cell.border.outline,
                                diagonalUp=template_cell.border.diagonalUp,
                                diagonalDown=template_cell.border.diagonalDown
                            )

                        if hasattr(template_cell, 'alignment') and template_cell.alignment:
                            new_cell.alignment = Alignment(
                                horizontal=template_cell.alignment.horizontal,
                                vertical=template_cell.alignment.vertical,
                                text_rotation=template_cell.alignment.text_rotation,
                                wrap_text=template_cell.alignment.wrap_text,
                                shrink_to_fit=template_cell.alignment.shrink_to_fit,
                                indent=template_cell.alignment.indent
                            )

                        if hasattr(template_cell, 'number_format') and template_cell.number_format:
                            new_cell.number_format = template_cell.number_format

                        # 🔧 设置统一的白色背景（而不是复制模板的背景色）
                        # 使用白色背景确保所有产品行都有一致的外观
                        white_fill = PatternFill(
                            fill_type="solid",
                            start_color="FFFFFFFF",
                            end_color="FFFFFFFF"
                        )
                        new_cell.fill = white_fill

                        # 恢复单元格的值
                        if current_value is not None:
                            new_cell.value = current_value

                    except Exception as e:
                        logger.warning(f"⚠️  为{col}{new_row}应用样式失败: {e}")
                        # 继续处理其他单元格，不中断整个过程

        # 5. 更新总计公式范围（总计区域已自动下移）
        last_product_row = first_product_row + product_count - 1

        # 🎯 计算总计区域的新位置（考虑插入的行数）
        if product_count > template_product_rows:
            additional_rows = product_count - template_product_rows
            new_totals_start_row = TOTALS_START_ROW + additional_rows
        else:
            new_totals_start_row = TOTALS_START_ROW

        subtotal_row = new_totals_start_row      # Sub Total行
        tax_row = subtotal_row + 1               # Tax行
        grand_total_row = tax_row + 1            # Grand Total行

        # 更新Sub Total公式
        subtotal_formula = f"=SUM(L{first_product_row}:L{last_product_row})"
        safe_write_cell(worksheet, f'L{subtotal_row}', subtotal_formula)

        # Tax计算（假设8%税率）
        tax_formula = f"=L{subtotal_row}*0.08"
        safe_write_cell(worksheet, f'L{tax_row}', tax_formula)

        # Grand Total公式
        grand_total_formula = f"=L{subtotal_row}+L{tax_row}"
        safe_write_cell(worksheet, f'L{grand_total_row}', grand_total_formula)

        logger.info(f"📊 更新总计公式范围: L{first_product_row}:L{last_product_row}")
        logger.info(f"🎯 总计区域位置: Sub Total={subtotal_row}, Tax={tax_row}, Grand Total={grand_total_row}")

        # 6. 强制重新计算所有公式
        workbook.calculation.calcMode = 'auto'
        workbook.calculation.fullCalcOnLoad = True

        # 6. 保存修改后的工作簿到内存
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        logger.info("✅ Excel文件生成完成")
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Excel生成失败: {str(e)}")
        raise Exception(f"Excel生成失败: {str(e)}")

@router.post("/generate-purchase-order")
async def generate_purchase_order_excel(
    request: PurchaseOrderRequest,
    db: Session = Depends(deps.get_db)
):
    """
    生成采购订单Excel文件
    """
    try:
        logger.info(f"开始生成供应商 {request.supplier_name} 的采购订单Excel")

        # 获取供应商信息
        supplier = db.query(Supplier).filter(Supplier.id == request.supplier_id).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="供应商不存在")

        # 生成Excel文件 - 使用新的生成器
        excel_content = create_purchase_order_excel_new(request, supplier, db)

        # 🔧 生成文件名：询价单_{supplier_name}_{date}.xlsx
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"询价单_{request.supplier_name}_{current_date}.xlsx"

        logger.info(f"Excel文件生成成功: {filename}")

        # 🔧 修复文件名编码问题：使用URL编码处理日文字符
        encoded_filename = urllib.parse.quote(filename, safe='')

        # 返回Excel文件
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成采购订单Excel失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"生成Excel失败: {str(e)}")


@router.post("/preview-purchase-order", response_model=ExcelPreviewData)
async def preview_purchase_order_excel(
    request: PurchaseOrderRequest,
    db: Session = Depends(deps.get_db)
):
    """
    预览采购订单Excel数据（不生成文件）
    """
    try:
        logger.info(f"开始预览供应商 {request.supplier_name} 的采购订单数据")

        # 获取供应商信息
        supplier = db.query(Supplier).filter(Supplier.id == request.supplier_id).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="供应商不存在")

        # 构建预览数据
        current_date = datetime.now()

        # 供应商信息
        supplier_info = {
            "name": supplier.name or "",
            "phone": supplier.phone or "",
            "contact": supplier.contact or "",
            "email": supplier.email or ""
        }

        # 订单信息
        order_info = {
            "date": current_date.strftime('%Y-%m-%d'),
            "invoice_number": request.invoice_number or f"{current_date.strftime('%Y%m%d')}-02 ML",
            "voyage_number": request.voyage_number or "ML-1017"
        }

        # 交货信息
        delivery_date_formatted = request.delivery_date
        if '/' in request.delivery_date:
            try:
                date_parts = request.delivery_date.split('/')
                if len(date_parts) == 3:
                    year, month, day = date_parts
                    delivery_date_formatted = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except:
                delivery_date_formatted = request.delivery_date

        delivery_info = {
            "delivery_date": delivery_date_formatted,
            "delivery_address": request.delivery_address or ""
        }

        # 计算总计
        subtotal = sum(product.amount for product in request.products)
        tax = subtotal * 0.08  # 8%税率
        total = subtotal + tax

        totals = {
            "subtotal": subtotal,
            "tax": tax,
            "total": total,
            "tax_rate": 0.08
        }

        preview_data = ExcelPreviewData(
            supplier_info=supplier_info,
            order_info=order_info,
            delivery_info=delivery_info,
            products=request.products,
            totals=totals
        )

        logger.info(f"预览数据生成成功")
        return preview_data

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"预览采购订单数据失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"预览失败: {str(e)}")


class UpdateAndGenerateRequest(BaseModel):
    """更新并生成请求"""
    update_request: ExcelUpdateRequest
    original_request: PurchaseOrderRequest

@router.post("/update-and-generate")
async def update_and_generate_excel(
    request: UpdateAndGenerateRequest,
    db: Session = Depends(deps.get_db)
):
    """
    根据用户修改更新并生成Excel文件
    """
    try:
        logger.info(f"开始根据用户修改生成Excel")

        update_request = request.update_request
        original_request = request.original_request

        logger.info(f"🔍 更新请求数据: {update_request}")
        logger.info(f"🔍 原始请求数据: {original_request}")

        # 获取供应商信息
        supplier = db.query(Supplier).filter(Supplier.id == original_request.supplier_id).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="供应商不存在")

        # 确保产品数据中的amount字段正确计算
        updated_products = []
        for product in update_request.products:
            # 重新计算amount字段
            amount = product.quantity * product.unit_price
            updated_product = ProductItem(
                po_number=product.po_number,
                product_code=product.product_code,
                product_name_en=product.product_name_en,
                product_name_jp=product.product_name_jp,
                pack_size=product.pack_size,
                quantity=product.quantity,
                unit=product.unit,
                unit_price=product.unit_price,
                amount=amount,  # 重新计算
                currency=product.currency
            )
            updated_products.append(updated_product)

        logger.info(f"🔧 重新计算了 {len(updated_products)} 个产品的金额")

        # 使用更新后的数据创建新的请求
        updated_request = PurchaseOrderRequest(
            supplier_id=original_request.supplier_id,
            supplier_name=update_request.supplier_info.get("name", original_request.supplier_name),
            products=updated_products,
            delivery_date=update_request.delivery_info.get("delivery_date", original_request.delivery_date),
            delivery_address=update_request.delivery_info.get("delivery_address", original_request.delivery_address),
            total_amount=sum(product.amount for product in updated_products),
            currency=original_request.currency,
            invoice_number=update_request.order_info.get("invoice_number", original_request.invoice_number),
            voyage_number=update_request.order_info.get("voyage_number", original_request.voyage_number)
        )

        logger.info(f"✅ 创建更新后的请求，总金额: {updated_request.total_amount}")

        # 生成Excel文件 - 使用新的生成器
        excel_content = create_purchase_order_excel_new(updated_request, supplier, db)

        # 生成文件名
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"询价单_{original_request.supplier_name}_{current_date}.xlsx"

        logger.info(f"更新后的Excel文件生成成功: {filename}")

        # 🔧 修复文件名编码问题：使用URL编码处理日文字符
        encoded_filename = urllib.parse.quote(filename, safe='')

        # 返回Excel文件
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新并生成Excel失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"更新生成失败: {str(e)}")


# PDF功能暂时移除，正在开发中





@router.post("/generate-purchase-order-pdf")
async def generate_purchase_order_pdf(
    request: PurchaseOrderRequest,
    db: Session = Depends(deps.get_db)
):
    """
    生成采购订单PDF文件 - 功能开发中
    """
    raise HTTPException(status_code=501, detail="PDF下载功能正在开发中，请使用Excel下载功能")
