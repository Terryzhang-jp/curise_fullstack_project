#!/usr/bin/env python3
"""
测试Excel模板动态扩展功能
验证当产品数量超过模板预设行数时，能否正确扩展模板
"""

import sys
import os
import json
import requests
from datetime import datetime

# 添加项目根目录到Python路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def test_excel_dynamic_expansion():
    """测试Excel动态扩展功能"""
    
    # 后端API地址
    base_url = "http://localhost:8000"
    
    print("🧪 开始测试Excel模板动态扩展功能...")
    
    # 创建测试数据：15个产品（超过模板的10行限制）
    test_products = []
    total_amount = 0
    for i in range(1, 16):  # 创建15个产品
        quantity = i * 10
        unit_price = 100 + i
        amount = quantity * unit_price
        total_amount += amount

        product = {
            "po_number": f"PO{i:08d}",
            "product_code": f"PRD{i:06d}",
            "product_name_en": f"Test Product {i}",
            "product_name_jp": f"テスト商品 {i}",
            "description": f"Test description for product {i}",
            "quantity": quantity,
            "unit": "CT",
            "unit_price": unit_price,
            "amount": amount,
            "currency": "JPY"
        }
        test_products.append(product)

    # 创建询价请求数据
    request_data = {
        "supplier_id": 1,  # 假设供应商ID为1
        "supplier_name": "测试供应商",
        "delivery_date": "2024-02-15",
        "delivery_address": "横浜港",
        "total_amount": total_amount,
        "currency": "JPY",
        "invoice_number": "INV-TEST-001",
        "voyage_number": "VOY-TEST-001",
        "products": test_products
    }
    
    print(f"📊 测试数据：{len(test_products)}个产品（超过模板10行限制）")
    
    try:
        # 发送请求到Excel生成API
        response = requests.post(
            f"{base_url}/api/v2/excel/generate-purchase-order",
            json=request_data,
            headers={"Content-Type": "application/json"}
        )
        
        if response.status_code == 200:
            print("✅ Excel生成成功！")
            
            # 保存生成的Excel文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_dynamic_expansion_{timestamp}.xlsx"
            filepath = os.path.join(os.path.dirname(__file__), filename)
            
            with open(filepath, 'wb') as f:
                f.write(response.content)
            
            print(f"📁 Excel文件已保存: {filepath}")
            print(f"📏 文件大小: {len(response.content)} bytes")
            
            # 验证文件是否可以正常打开
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath)
                ws = wb.active
                
                # 检查产品数据是否正确填充
                product_rows_found = 0
                for row in range(22, 50):  # 检查第22-49行
                    cell_value = ws[f'A{row}'].value
                    if cell_value is not None and str(cell_value).isdigit():
                        product_rows_found += 1
                
                print(f"📋 检测到产品行数: {product_rows_found}")
                
                if product_rows_found >= 15:
                    print("🎉 动态扩展功能测试成功！")
                    print("✅ 所有15个产品都正确填充到Excel中")
                else:
                    print(f"❌ 动态扩展功能测试失败！只找到{product_rows_found}个产品")
                
                wb.close()
                
            except Exception as e:
                print(f"❌ Excel文件验证失败: {e}")
                
        else:
            print(f"❌ Excel生成失败！状态码: {response.status_code}")
            print(f"错误信息: {response.text}")
            
    except requests.exceptions.ConnectionError:
        print("❌ 无法连接到后端服务，请确保后端正在运行 (python main.py)")
    except Exception as e:
        print(f"❌ 测试过程中发生错误: {e}")

if __name__ == "__main__":
    test_excel_dynamic_expansion()
