#!/usr/bin/env python3
"""
验证Excel动态扩展功能的详细脚本
检查生成的Excel文件是否正确扩展了产品行
"""

import sys
import os
from openpyxl import load_workbook
import glob

def verify_excel_expansion():
    """验证Excel动态扩展功能"""
    
    print("🔍 开始验证Excel动态扩展功能...")
    
    # 查找最新的测试文件
    test_files = glob.glob("test/test_dynamic_expansion_*.xlsx")
    if not test_files:
        print("❌ 未找到测试生成的Excel文件")
        return
    
    latest_file = max(test_files, key=os.path.getctime)
    print(f"📁 验证文件: {latest_file}")
    
    try:
        # 打开Excel文件
        wb = load_workbook(latest_file)
        ws = wb.active
        
        print(f"📊 工作表名称: {ws.title}")
        print(f"📏 工作表尺寸: {ws.max_row} 行 x {ws.max_column} 列")
        print()
        
        # 检查产品数据区域
        print("🔍 检查产品数据区域:")
        print("=" * 60)
        
        product_count = 0
        for row in range(22, 50):  # 检查第22-49行
            # 检查NO.列（A列）
            no_value = ws[f'A{row}'].value
            # 检查产品代码列（C列）
            code_value = ws[f'C{row}'].value
            # 检查产品名称列（D列）
            name_value = ws[f'D{row}'].value
            # 检查数量列（H列）
            quantity_value = ws[f'H{row}'].value
            # 检查金额公式列（L列）
            amount_value = ws[f'L{row}'].value
            
            if no_value is not None and str(no_value).isdigit():
                product_count += 1
                print(f"第{row}行: NO.={no_value}, 代码={code_value}, 名称={name_value}, 数量={quantity_value}")
                
                # 检查金额公式
                if hasattr(amount_value, 'value') and str(amount_value.value).startswith('='):
                    print(f"        ✅ 金额公式: {amount_value.value}")
                elif isinstance(amount_value, str) and amount_value.startswith('='):
                    print(f"        ✅ 金额公式: {amount_value}")
                else:
                    print(f"        ⚠️  金额值: {amount_value} (可能不是公式)")
                print()
        
        print("=" * 60)
        print(f"📋 总计检测到产品行数: {product_count}")
        
        # 验证结果
        expected_products = 15
        if product_count >= expected_products:
            print(f"🎉 动态扩展功能验证成功！")
            print(f"✅ 期望产品数: {expected_products}")
            print(f"✅ 实际产品数: {product_count}")
            print(f"✅ 扩展行数: {max(0, product_count - 10)}")
        else:
            print(f"❌ 动态扩展功能验证失败！")
            print(f"❌ 期望产品数: {expected_products}")
            print(f"❌ 实际产品数: {product_count}")
        
        # 检查模板结构是否保持完整
        print("\n🏗️ 检查模板结构完整性:")
        print("=" * 60)
        
        # 检查标题行
        title_cell = ws['A1'].value
        print(f"标题 (A1): {title_cell}")
        
        # 检查供应商信息区域
        supplier_label = ws['A5'].value
        supplier_value = ws['B5'].value
        print(f"供应商标签 (A5): {supplier_label}")
        print(f"供应商值 (B5): {supplier_value}")
        
        # 检查表头
        header_row = 21
        headers = []
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            header_value = ws[f'{col}{header_row}'].value
            headers.append(f"{col}:{header_value}")
        
        print(f"表头 (第{header_row}行): {', '.join(headers)}")
        
        wb.close()
        
        print("\n✅ Excel文件验证完成！")
        
    except Exception as e:
        print(f"❌ Excel文件验证失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verify_excel_expansion()
